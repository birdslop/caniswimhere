#!/usr/bin/env python3
"""
Ingest storm overflow assets and annual returns for ALL water companies
from the EDM 2024 Storm Overflow Annual Return xlsx.

Extends edm_thames_2024_fetcher.py to iterate every company sheet.
Idempotent: safe to re-run (ON CONFLICT DO NOTHING on natural keys).
"""
import os
import sys
import openpyxl
import psycopg2
from psycopg2.extras import Json
from datetime import datetime, timedelta

_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(
    _REPO_ROOT,
    "data", "edm", "EDM_2024_Storm_Overflow_Annual_Return",
    "EDM 2024 Storm Overflow Annual Return - all water and sewerage companies.xlsx",
)

# All company sheets in the workbook (2024 annual return).
COMPANY_SHEETS = [
    "Anglian Water 2024",
    "DCWW 2024",
    "Northumbrian Water 2024",
    "Severn Trent 2024",
    "South West Water 2024",
    "Southern Water 2024",
    "Thames Water 2024",
    "United Utilities 2024",
    "Wessex Water 2024",
    "Yorkshire Water 2024",
]


def json_safe(value):
    """Convert Excel / Python objects into JSON-safe representations."""
    if isinstance(value, timedelta):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def make_record_safe(record: dict) -> dict:
    return {k: json_safe(v) for k, v in record.items()}


def safe_int(value):
    """Coerce to int if possible, else None. Preserves raw value in raw_metadata."""
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    try:
        return int(str(value).strip())
    except (ValueError, TypeError):
        return None


def safe_numeric(value):
    """Coerce to float if possible, else None."""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def ingest_sheet(cur, ws, sheet_name: str, source_id) -> tuple[int, int, int]:
    """Ingest a single company sheet. Returns (overflows_inserted, returns_inserted, skipped)."""
    headers = [cell.value for cell in ws[2]]
    if not headers or headers[0] is None:
        print(f"  WARNING: sheet {sheet_name!r} has no headers in row 2, skipping.")
        return 0, 0, 0

    inserted_overflows = 0
    inserted_returns = 0
    skipped = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        record = dict(zip(headers, row))
        record_safe = make_record_safe(record)

        unique_id = record_safe.get("Unique ID")
        if not unique_id:
            skipped += 1
            continue

        # Insert overflow asset
        cur.execute("""
            INSERT INTO overflows (
                unique_id,
                water_company_name,
                site_name_ea,
                site_name_wasc,
                ea_permit_reference,
                activity_reference,
                asset_type,
                outlet_discharge_ngr,
                wfd_waterbody_id,
                wfd_catchment_name,
                receiving_water_name,
                source_id,
                raw_metadata
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (unique_id) DO NOTHING
        """, (
            unique_id,
            record_safe.get("Water Company Name"),
            record_safe.get("Site Name\n(EA Consents Database)"),
            record_safe.get("Site Name\n(WaSC operational)\n[optional]"),
            record_safe.get("EA Permit Reference\n(EA Consents Database)"),
            record_safe.get("Activity Reference on Permit"),
            record_safe.get("Storm Discharge Asset Type"),
            record_safe.get("Outlet Discharge NGR\n(EA Consents Database)"),
            record_safe.get("WFD Waterbody ID (Cycle 3)\n(discharge outlet)"),
            record_safe.get("WFD Waterbody Catchment Name (Cycle 3)\n(discharge outlet)"),
            record_safe.get("Receiving Water / Environment (common name)\n(EA Consents Database)"),
            source_id,
            Json(record_safe),
        ))

        if cur.rowcount == 1:
            inserted_overflows += 1

        # Insert annual return (2024)
        cur.execute("""
            INSERT INTO overflow_annual_returns (
                unique_id,
                report_year,
                data_start_year,
                edm_operational_pct,
                total_duration_text,
                counted_spills,
                long_term_avg_spill_count,
                source_id,
                raw_metadata
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (unique_id, report_year) DO NOTHING
        """, (
            unique_id,
            2024,
            safe_int(record_safe.get("Data start - calendar year")),
            safe_numeric(record_safe.get("EDM Operation -\n% of reporting period EDM operational")),
            str(record_safe.get(
                "Total Duration (hh:mm:ss) all spills prior to processing through 12-24h count method"
            ) or "") or None,
            safe_int(record_safe.get("Counted spills using 12-24h count method")),
            safe_numeric(record_safe.get("Long-term average spill count")),
            source_id,
            Json(record_safe),
        ))

        if cur.rowcount == 1:
            inserted_returns += 1

    return inserted_overflows, inserted_returns, skipped


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    available_sheets = set(wb.sheetnames)

    conn = psycopg2.connect(dbname="water_quality")
    cur = conn.cursor()

    # Single provenance row for this full national ingest run.
    cur.execute("""
        INSERT INTO sources (
            provider,
            dataset_name,
            source_url,
            license,
            fetched_at,
            raw_metadata
        )
        VALUES (%s, %s, %s, %s, %s, %s)
        RETURNING source_id
    """, (
        "Environment Agency",
        "EDM Storm Overflow Annual Return 2024 (All Companies)",
        "https://environment.data.gov.uk/datasets/storm-overflows",
        "Open Government Licence v3.0",
        datetime.utcnow(),
        Json({"sheets": COMPANY_SHEETS}),
    ))
    source_id = cur.fetchone()[0]

    total_overflows = 0
    total_returns = 0
    total_skipped = 0

    for sheet_name in COMPANY_SHEETS:
        if sheet_name not in available_sheets:
            print(f"  WARNING: sheet {sheet_name!r} not found in workbook, skipping.")
            continue

        ws = wb[sheet_name]
        ov, ret, sk = ingest_sheet(cur, ws, sheet_name, source_id)
        total_overflows += ov
        total_returns += ret
        total_skipped += sk
        print(f"  {sheet_name}: {ov} overflows, {ret} returns, {sk} skipped")

    conn.commit()
    cur.close()
    conn.close()
    wb.close()

    print(f"\nTotal: {total_overflows} overflows inserted, {total_returns} returns inserted, {total_skipped} skipped.")


if __name__ == "__main__":
    main()
