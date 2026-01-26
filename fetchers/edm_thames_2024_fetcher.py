import openpyxl
import psycopg2
from psycopg2.extras import Json
from datetime import datetime, timedelta

XLSX_PATH = (
    "data/edm/EDM_2024_Storm_Overflow_Annual_Return/"
    "EDM 2024 Storm Overflow Annual Return - all water and sewerage companies.xlsx"
)

SHEET_NAME = "Thames Water 2024"

def json_safe(value):
    """
    Convert Excel / Python objects into JSON-safe representations.
    """
    if isinstance(value, timedelta):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    return value

def make_record_safe(record: dict) -> dict:
    return {k: json_safe(v) for k, v in record.items()}

def main():
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]

    conn = psycopg2.connect(dbname="water_quality")
    cur = conn.cursor()

    # provenance
    cur.execute("""
        INSERT INTO sources (
            provider,
            dataset_name,
            source_url,
            license,
            fetched_at,
            raw_metadata
        )
        VALUES (%s, %s, %s, %s, %s, %s)
        RETURNING source_id
    """, (
        "Environment Agency",
        "EDM Storm Overflow Annual Return 2024 (Thames Water)",
        "https://environment.data.gov.uk/datasets/storm-overflows",
        "Open Government Licence v3.0",
        datetime.utcnow(),
        Json({"sheet": SHEET_NAME})
    ))
    source_id = cur.fetchone()[0]

    inserted_overflows = 0
    inserted_returns = 0

    headers = [cell.value for cell in ws[2]]

    for row in ws.iter_rows(min_row=3, values_only=True):
        record = dict(zip(headers, row))
        record_safe = make_record_safe(record)

        unique_id = record_safe.get("Unique ID")
        if not unique_id:
            continue

        # Insert overflow asset
        cur.execute("""
            INSERT INTO overflows (
                unique_id,
                water_company_name,
                site_name_ea,
                site_name_wasc,
                ea_permit_reference,
                activity_reference,
                asset_type,
                outlet_discharge_ngr,
                wfd_waterbody_id,
                wfd_catchment_name,
                receiving_water_name,
                source_id,
                raw_metadata
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (unique_id) DO NOTHING
        """, (
            unique_id,
            record_safe.get("Water Company Name"),
            record_safe.get("Site Name\n(EA Consents Database)"),
            record_safe.get("Site Name\n(WaSC operational)\n[optional]"),
            record_safe.get("EA Permit Reference\n(EA Consents Database)"),
            record_safe.get("Activity Reference on Permit"),
            record_safe.get("Storm Discharge Asset Type"),
            record_safe.get("Outlet Discharge NGR\n(EA Consents Database)"),
            record_safe.get("WFD Waterbody ID (Cycle 3)\n(discharge outlet)"),
            record_safe.get("WFD Waterbody Catchment Name (Cycle 3)\n(discharge outlet)"),
            record_safe.get("Receiving Water / Environment (common name)\n(EA Consents Database)"),
            source_id,
            Json(record_safe)
        ))

        if cur.rowcount == 1:
            inserted_overflows += 1

        # Insert annual return (2024)
        cur.execute("""
            INSERT INTO overflow_annual_returns (
                unique_id,
                report_year,
                data_start_year,
                edm_operational_pct,
                total_duration_text,
                counted_spills,
                long_term_avg_spill_count,
                source_id,
                raw_metadata
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (unique_id, report_year) DO NOTHING
        """, (
            unique_id,
            2024,
            record_safe.get("Data start - calendar year"),
            record_safe.get("EDM Operation -\n% of reporting period EDM operational"),
            record_safe.get(
                "Total Duration (hh:mm:ss) all spills prior to processing through 12-24h count method"
            ),
            record_safe.get("Counted spills using 12-24h count method"),
            record_safe.get("Long-term average spill count"),
            source_id,
            Json(record_safe)
        ))

        if cur.rowcount == 1:
            inserted_returns += 1

    conn.commit()
    cur.close()
    conn.close()

    print(f"Inserted {inserted_overflows} overflow assets.")
    print(f"Inserted {inserted_returns} annual return rows.")

if __name__ == "__main__":
    main()
